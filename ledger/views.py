from datetime import date
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from .forms import BudgetForm, CategoryForm, ExpenseForm, IncomeForm, RegisterForm, ReportFilterForm
from .models import Budget, Category, Expense, Income

def register(request):
    if request.user.is_authenticated: return redirect("dashboard")
    form = RegisterForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        Category.objects.bulk_create([
            Category(user=user, name="Salary", type=Category.INCOME),
            Category(user=user, name="Other income", type=Category.INCOME),
            Category(user=user, name="Food", type=Category.EXPENSE),
            Category(user=user, name="Transport", type=Category.EXPENSE),
            Category(user=user, name="Bills", type=Category.EXPENSE),
            Category(user=user, name="Other expenses", type=Category.EXPENSE),
        ])
        login(request, user); return redirect("dashboard")
    return render(request, "ledger/register.html", {"form": form})

@login_required
def dashboard(request):
    today = date.today(); month_start = today.replace(day=1)
    income = Income.objects.filter(user=request.user, date__gte=month_start).aggregate(total=Sum("amount"))["total"] or Decimal("0")
    expense = Expense.objects.filter(user=request.user, date__gte=month_start).aggregate(total=Sum("amount"))["total"] or Decimal("0")
    return render(request, "ledger/dashboard.html", {"income": income, "expense": expense, "balance": income-expense, "recent": list(Income.objects.filter(user=request.user)[:5]) + list(Expense.objects.filter(user=request.user)[:5])})

def form_page(request, form_class, template, title, instance=None):
    form = form_class(request.POST or None, instance=instance, user=request.user)
    # Attach ownership before validation so Django can validate model-level
    # uniqueness rules and report errors in the form instead of returning 500.
    if instance is None:
        form.instance.user = request.user
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.save(); messages.success(request, f"{title} saved."); return redirect(template.split("_")[0] + "_list")
    return render(request, "ledger/form.html", {"form": form, "title": title})

@login_required
def categories(request):
    if request.method == "POST":
        form = CategoryForm(request.POST, user=request.user)
        form.instance.user = request.user
        if form.is_valid(): form.save(); messages.success(request, "Category created."); return redirect("categories")
    else: form = CategoryForm(user=request.user)
    return render(request, "ledger/categories.html", {"form": form, "categories": Category.objects.filter(user=request.user)})

@login_required
def income_list(request): return render(request, "ledger/transactions.html", {"items": Income.objects.filter(user=request.user), "kind": "Income", "create_url": "income_create", "edit_url": "income_edit", "delete_url": "income_delete"})
@login_required
def expense_list(request): return render(request, "ledger/transactions.html", {"items": Expense.objects.filter(user=request.user), "kind": "Expense", "create_url": "expense_create", "edit_url": "expense_edit", "delete_url": "expense_delete"})
@login_required
def income_create(request): return form_page(request, IncomeForm, "income_create", "Income")
@login_required
def expense_create(request): return form_page(request, ExpenseForm, "expense_create", "Expense")
@login_required
def income_edit(request, pk): return form_page(request, IncomeForm, "income_edit", "Income", get_object_or_404(Income, pk=pk, user=request.user))
@login_required
def expense_edit(request, pk): return form_page(request, ExpenseForm, "expense_edit", "Expense", get_object_or_404(Expense, pk=pk, user=request.user))

def delete_item(request, model, pk, destination):
    item = get_object_or_404(model, pk=pk, user=request.user)
    if request.method == "POST": item.delete(); messages.success(request, "Record deleted."); return redirect(destination)
    return render(request, "ledger/confirm_delete.html", {"item": item})
@login_required
def income_delete(request, pk): return delete_item(request, Income, pk, "income_list")
@login_required
def expense_delete(request, pk): return delete_item(request, Expense, pk, "expense_list")

@login_required
def budgets(request):
    form = BudgetForm(request.POST or None, user=request.user)
    form.instance.user = request.user
    if request.method == "POST" and form.is_valid(): form.save(); messages.success(request, "Budget saved."); return redirect("budgets")
    rows=[]
    for budget in Budget.objects.filter(user=request.user).select_related("category"):
        spent = Expense.objects.filter(user=request.user, category=budget.category, date__year=budget.month.year, date__month=budget.month.month).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        rows.append((budget, spent, min(int((spent / budget.amount_limit) * 100), 100)))
    return render(request, "ledger/budgets.html", {"form": form, "rows": rows})

def report_data(user, start_date=None, end_date=None):
    expenses = Expense.objects.filter(user=user)
    income = Income.objects.filter(user=user)
    if start_date:
        expenses = expenses.filter(date__gte=start_date)
        income = income.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)
        income = income.filter(date__lte=end_date)
    context = {"start_date": start_date, "end_date": end_date,
               "income_total": income.aggregate(total=Sum("amount"))["total"] or 0, "expense_total": expenses.aggregate(total=Sum("amount"))["total"] or 0,
               "by_category": expenses.values("category__name").annotate(total=Sum("amount")).order_by("-total"),
               "monthly_income": income.annotate(month=TruncMonth("date")).values("month").annotate(total=Sum("amount")).order_by("month"),
               "monthly_expense": expenses.annotate(month=TruncMonth("date")).values("month").annotate(total=Sum("amount")).order_by("month")}
    context["balance"] = context["income_total"] - context["expense_total"]
    return context

@login_required
def reports(request):
    filter_form = ReportFilterForm(request.GET)
    context = report_data(request.user, **filter_form.cleaned_data) if filter_form.is_valid() else report_data(request.user)
    context["filter_form"] = filter_form
    return render(request, "ledger/reports.html", context)


def validated_report_data(request):
    filter_form = ReportFilterForm(request.GET)
    if not filter_form.is_valid():
        return None, HttpResponse("Invalid report date range.", status=400)
    return report_data(request.user, **filter_form.cleaned_data), None

@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    data, error_response = validated_report_data(request)
    if error_response:
        return error_response
    workbook = Workbook(); summary = workbook.active; summary.title = "Summary"
    summary.append(["MoneyTrack Financial Report"]); summary["A1"].font = Font(bold=True, size=16)
    summary.append([]); summary.append(["Metric", "Amount (NPR)"])
    for cell in summary[3]: cell.font = Font(bold=True)
    summary.append(["Total income", data["income_total"]]); summary.append(["Total expenses", data["expense_total"]]); summary.append(["Net balance", data["balance"]])
    categories = workbook.create_sheet("Expenses by Category"); categories.append(["Category", "Amount (NPR)"])
    for cell in categories[1]: cell.font = Font(bold=True)
    for row in data["by_category"]: categories.append([row["category__name"], row["total"]])
    transactions = workbook.create_sheet("Transactions"); transactions.append(["Type", "Date", "Category", "Amount", "Currency", "Description", "Payment method"])
    for cell in transactions[1]: cell.font = Font(bold=True)
    income_items = Income.objects.filter(user=request.user).select_related("category")
    expense_items = Expense.objects.filter(user=request.user).select_related("category")
    if request.GET.get("start_date"):
        income_items = income_items.filter(date__gte=data["start_date"])
        expense_items = expense_items.filter(date__gte=data["start_date"])
    if request.GET.get("end_date"):
        income_items = income_items.filter(date__lte=data["end_date"])
        expense_items = expense_items.filter(date__lte=data["end_date"])
    for item in income_items: transactions.append(["Income", item.date, item.category.name, item.amount, item.currency, item.description, ""])
    for item in expense_items: transactions.append(["Expense", item.date, item.category.name, item.amount, item.currency, item.description, item.get_payment_method_display()])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="moneytrack-report.xlsx"'; workbook.save(response)
    return response

@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Spacer, Paragraph, Table, TableStyle
    data, error_response = validated_report_data(request)
    if error_response:
        return error_response
    response = HttpResponse(content_type="application/pdf"); response["Content-Disposition"] = 'attachment; filename="moneytrack-report.pdf"'
    styles = getSampleStyleSheet(); story = [Paragraph("MoneyTrack Financial Report", styles["Title"]), Spacer(1, 14)]
    tables = [("Summary", [["Metric", "Amount (NPR)"], ["Total income", str(data["income_total"])], ["Total expenses", str(data["expense_total"])], ["Net balance", str(data["balance"])]]), ("Expenses by category", [["Expense category", "Amount (NPR)"]] + [[row["category__name"], str(row["total"])] for row in data["by_category"]])]
    for heading, rows in tables:
        story += [Paragraph(heading, styles["Heading2"]), Table(rows, colWidths=[260, 180], style=TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0d6efd")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .4, colors.lightgrey), ("PADDING", (0,0), (-1,-1), 7)])), Spacer(1, 16)]
    SimpleDocTemplate(response, pagesize=A4, rightMargin=55, leftMargin=55, topMargin=55, bottomMargin=55).build(story)
    return response
